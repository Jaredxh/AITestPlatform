"""测试用例 Excel 模板：导入 / 导出 / 模板生成。

设计目标（与产品验收口径对齐）：

1. **同一份模板既能导出也能导入**：用户可以"导出 → 改 → 再导入覆盖"。
2. **新增 vs 更新由"用例编号"列决定**：
   - 用例编号为空 → 新增
   - 用例编号填了，但项目里查不到 → 当作新增（用户可以提前规划编号但暂时跑空）
   - 用例编号填了且项目里查得到 → 按编号匹配做"覆盖式更新"
3. **模块路径用 ``/`` 分层**：导入时找不到的层级会自动创建，让用户可以一份
   表格同时拉一批新模块和它们底下的用例。
4. **步骤分两列各自多行**：``步骤动作`` 与 ``预期结果`` 各自一个 cell，cell 内
   按行对齐，第 N 行就是第 N 个步骤的动作 / 预期；缺少的一边补空。这是
   Excel 里写人能看的"两列对照表"的最佳折中——比让人在一个 cell 里敲
   ``1|click|ok / 2|input|ok`` 这种密码学符号清晰得多。

不在本模块的职责：
- DB 写入：本模块只产出"中间结构 ``ParsedRow``"，由 service 层接管。
- 权限校验：交给 router 的 ``require_permission``。
- 数据库 case_no 分配：只在 service 层调 ``_allocate_case_no``。
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.modules.testcases.models import Testcase, TestcaseModule
from app.modules.testcases.schemas import format_case_display_id

# ─── 常量：列定义、值映射 ─────────────────────────────────────────────

SHEET_NAME = "测试用例"
INSTRUCTIONS_SHEET_NAME = "字段说明"

# 列顺序 = 模板表头顺序。每一项 (header, key, width)。
COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("用例编号", "display_id", 14),
    ("模块路径", "module_path", 24),
    ("用例标题", "title", 40),
    ("优先级", "priority", 10),
    ("状态", "status", 10),
    ("执行结果", "exec_result", 12),
    ("前置条件", "precondition", 30),
    ("步骤动作", "steps_action", 40),
    ("预期结果", "steps_expected", 40),
)

# header → key 的反查表（小写归一化），允许用户改个表头大小写不挂掉
HEADER_ALIASES: dict[str, str] = {h.strip().lower(): k for h, k, _ in COLUMNS}
# 兼容用户把"步骤"作为 steps_action 的别名（导出后被 PM 习惯性改了表头）
HEADER_ALIASES.update({
    "步骤": "steps_action",
    "操作步骤": "steps_action",
    "测试步骤": "steps_action",
    "预期": "steps_expected",
    "预期值": "steps_expected",
    "id": "display_id",
    "用例 id": "display_id",
    "模块": "module_path",
    "标题": "title",
})

# 值的中英文双向映射。导出时优先写中文（人友好），导入时两种都接。
PRIORITY_DISPLAY: dict[str, str] = {"high": "高", "medium": "中", "low": "低"}
STATUS_DISPLAY: dict[str, str] = {"active": "有效", "draft": "草稿", "deprecated": "废弃"}
EXEC_RESULT_DISPLAY: dict[str, str] = {
    "not_run": "未执行",
    "passed": "通过",
    "failed": "失败",
    "blocked": "阻塞",
}

PRIORITY_LOOKUP: dict[str, str] = {
    **{v: k for k, v in PRIORITY_DISPLAY.items()},
    **{k: k for k in PRIORITY_DISPLAY},
    "h": "high", "m": "medium", "l": "low",
    "p0": "high", "p1": "high", "p2": "medium", "p3": "low",
}
STATUS_LOOKUP: dict[str, str] = {
    **{v: k for k, v in STATUS_DISPLAY.items()},
    **{k: k for k in STATUS_DISPLAY},
}
EXEC_RESULT_LOOKUP: dict[str, str] = {
    **{v: k for k, v in EXEC_RESULT_DISPLAY.items()},
    **{k: k for k in EXEC_RESULT_DISPLAY},
    "pass": "passed", "fail": "failed", "block": "blocked", "skip": "blocked",
}

# 模板里下拉枚举的备选项（写中文给最终用户看）
PRIORITY_OPTIONS = list(PRIORITY_DISPLAY.values())
STATUS_OPTIONS = list(STATUS_DISPLAY.values())
EXEC_RESULT_OPTIONS = list(EXEC_RESULT_DISPLAY.values())

# 用例编号正则：``TC-0001`` / ``tc-12``。导入时容错：纯数字也接。
DISPLAY_ID_RE = re.compile(r"^\s*TC[-_]?(\d+)\s*$", re.IGNORECASE)

# 单次导入行数硬上限——保护内存 + 避免误传全公司 Excel 把库写炸
EXCEL_IMPORT_MAX_ROWS = 5_000


# ─── 解析后的中间结构 ────────────────────────────────────────────────


@dataclass
class ParsedStep:
    step_number: int
    action: str
    expected_result: str | None


@dataclass
class ParsedRow:
    """一个 Excel 数据行的归一化结果。

    字段都用"业务语义"的名字，不带 Excel 行号——行号留在 ``row_no`` 里专门给
    报错回显。``case_no`` 仅当用户在表格里给了合法用例编号才被填上。
    """

    row_no: int
    title: str
    module_path: str | None = None
    case_no: int | None = None
    priority: str = "medium"
    status: str = "active"
    exec_result: str = "not_run"
    precondition: str | None = None
    steps: list[ParsedStep] = field(default_factory=list)


@dataclass
class ParseError:
    row: int
    message: str
    title: str | None = None


# ─── 模板生成 ───────────────────────────────────────────────────────


def build_template_workbook(*, sample: bool = True) -> bytes:
    """返回标准模板的 xlsx 字节流。

    ``sample=True`` 会在主表里塞两行示例（一行新增样例 + 一行更新样例），让
    第一次接触模板的人不必先去翻文档。导出时调 ``sample=False``。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _write_headers(ws)

    if sample:
        sample_rows = [
            {
                "display_id": "",
                "module_path": "登录模块/账号登录",
                "title": "示例：用正确账号登录系统",
                "priority": "高",
                "status": "有效",
                "exec_result": "未执行",
                "precondition": "已注册账号 admin/admin123",
                "steps_action": "打开登录页\n输入用户名密码\n点击登录按钮",
                "steps_expected": "登录页加载成功\n输入框可正常输入\n跳转至首页且右上角显示用户名",
            },
            {
                "display_id": "TC-0001",
                "module_path": "登录模块/账号登录",
                "title": "示例：编号已存在 → 覆盖原用例",
                "priority": "中",
                "status": "草稿",
                "exec_result": "未执行",
                "precondition": "",
                "steps_action": "步骤一动作\n步骤二动作",
                "steps_expected": "步骤一预期\n步骤二预期",
            },
        ]
        for offset, row in enumerate(sample_rows, start=2):
            for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
                ws.cell(row=offset, column=col_idx, value=row[key])
        # 示例行用浅灰背景区分一下，避免被当成"真数据"提交
        sample_fill = PatternFill("solid", fgColor="FFF8F4E5")
        for r in range(2, 2 + len(sample_rows)):
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=c).fill = sample_fill

    _apply_data_validations(ws, max_row=1000)
    _add_instructions_sheet(wb)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ─── 导出 ───────────────────────────────────────────────────────────


def build_export_workbook(
    testcases: Iterable[Testcase],
    module_path_resolver,
) -> bytes:
    """把一批用例 + 步骤渲染成 xlsx 字节流。

    ``module_path_resolver`` 是个函数：``module_id -> module_path``（``"/" 分隔``）。
    抽成参数是为了让 service 层一次性预算好整棵模块树的路径，避免每行都查 DB。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _write_headers(ws)

    for r, tc in enumerate(testcases, start=2):
        steps_sorted = sorted(tc.steps or [], key=lambda s: s.step_number)
        steps_action = "\n".join((s.action or "").strip() for s in steps_sorted)
        steps_expected = "\n".join((s.expected_result or "").strip() for s in steps_sorted)
        values = {
            "display_id": format_case_display_id(tc.case_no),
            "module_path": module_path_resolver(tc.module_id) or "",
            "title": tc.title,
            "priority": PRIORITY_DISPLAY.get(tc.priority, tc.priority),
            "status": STATUS_DISPLAY.get(tc.status, tc.status),
            "exec_result": EXEC_RESULT_DISPLAY.get(tc.exec_result, tc.exec_result),
            "precondition": tc.precondition or "",
            "steps_action": steps_action,
            "steps_expected": steps_expected,
        }
        for c, (_, key, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=c, value=values[key])

    # 导出文件也带下拉，让用户改完单元格后还有下拉值校验提示
    _apply_data_validations(ws, max_row=max(ws.max_row, 1000))
    _add_instructions_sheet(wb)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ─── 导入：解析 Excel → ParsedRow ────────────────────────────────────


def parse_workbook(raw: bytes) -> tuple[list[ParsedRow], list[ParseError]]:
    """把上传的 xlsx 字节流解析成 ParsedRow 列表 + 行级错误。

    设计原则：
    - 容错为主：能解析的都解析；遇到不认识的列名就忽略；空行跳过；
    - 行级错误聚合，不中断后续行；
    - 全局结构问题（无表头 / 文件不是 xlsx / 表头缺必需列 ``用例标题``）抛
      ``ValueError``——由 service 层翻成 ``AppException``。
    """
    try:
        # data_only=True 让公式被读成"上次保存的计算结果"，避免拿到 ``=A1+B1``
        # 这种字面值；导入测试用例不需要公式，结果即一切。
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"无法解析 Excel 文件：{exc}") from exc

    ws = _pick_data_sheet(wb)
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel 工作表为空，缺少表头") from exc

    header_to_index: dict[str, int] = {}
    for idx, raw_header in enumerate(header_row):
        if raw_header is None:
            continue
        norm = str(raw_header).strip().lower()
        key = HEADER_ALIASES.get(norm)
        if key and key not in header_to_index:
            header_to_index[key] = idx

    if "title" not in header_to_index:
        raise ValueError("表头缺少必需列：用例标题")

    parsed: list[ParsedRow] = []
    errors: list[ParseError] = []

    for excel_row_no, row_values in enumerate(rows_iter, start=2):
        if excel_row_no - 1 > EXCEL_IMPORT_MAX_ROWS:
            errors.append(ParseError(
                row=excel_row_no,
                message=f"超过单次导入上限 {EXCEL_IMPORT_MAX_ROWS} 行，后续行已忽略",
            ))
            break

        # 整行空白 → 跳过（Excel 末尾经常拖一堆空行）
        if not any(_is_meaningful(v) for v in row_values):
            continue

        cell = lambda key: _cell_str(row_values, header_to_index.get(key))  # noqa: E731

        title = cell("title")
        if not title:
            errors.append(ParseError(
                row=excel_row_no,
                message="缺少『用例标题』，本行已跳过",
            ))
            continue

        case_no_val: int | None = None
        display_raw = cell("display_id")
        if display_raw:
            case_no_val = _parse_display_id(display_raw)
            if case_no_val is None:
                errors.append(ParseError(
                    row=excel_row_no, title=title,
                    message=f"用例编号格式错误：{display_raw!r}（应为 TC-数字 或纯数字）",
                ))
                # 编号格式错误不致命：当作"新增"继续，且把错误抛给用户看
                case_no_val = None

        priority_raw = cell("priority")
        priority = _normalize_value(priority_raw, PRIORITY_LOOKUP, default="medium")
        if priority_raw and priority is None:
            errors.append(ParseError(
                row=excel_row_no, title=title,
                message=f"优先级取值非法：{priority_raw!r}（可选：高/中/低）",
            ))
            continue
        if priority is None:
            priority = "medium"

        status_raw = cell("status")
        status = _normalize_value(status_raw, STATUS_LOOKUP, default="active")
        if status_raw and status is None:
            errors.append(ParseError(
                row=excel_row_no, title=title,
                message=f"状态取值非法：{status_raw!r}（可选：有效/草稿/废弃）",
            ))
            continue
        if status is None:
            status = "active"

        exec_raw = cell("exec_result")
        exec_result = _normalize_value(exec_raw, EXEC_RESULT_LOOKUP, default="not_run")
        if exec_raw and exec_result is None:
            errors.append(ParseError(
                row=excel_row_no, title=title,
                message=f"执行结果取值非法：{exec_raw!r}（可选：未执行/通过/失败/阻塞）",
            ))
            continue
        if exec_result is None:
            exec_result = "not_run"

        steps = _parse_steps(cell("steps_action"), cell("steps_expected"))

        parsed.append(ParsedRow(
            row_no=excel_row_no,
            title=title,
            module_path=cell("module_path") or None,
            case_no=case_no_val,
            priority=priority,
            status=status,
            exec_result=exec_result,
            precondition=cell("precondition") or None,
            steps=steps,
        ))

    return parsed, errors


# ─── 模块路径工具：路径 → ``[name1, name2, ...]`` ─────────────────────


def split_module_path(path: str | None) -> list[str]:
    """``"a / b / c"`` → ``["a", "b", "c"]``；空 / None → ``[]``。"""
    if not path:
        return []
    parts = [p.strip() for p in path.split("/")]
    return [p for p in parts if p]


def join_module_path(names: Iterable[str]) -> str:
    return "/".join(names)


def build_module_path_index(
    modules: Iterable[TestcaseModule],
) -> tuple[dict[str | None, str], dict[tuple[str, ...], list[TestcaseModule]]]:
    """根据扁平的模块列表构造两个查询索引。

    返回：
    - ``id_to_path``: ``module_id → "登录/账号登录"``。``None`` 也加一项映射成
      ``""``，方便导出时无脑 lookup。
    - ``path_to_modules``: ``("登录", "账号登录") → [TestcaseModule, ...]``——
      key 是 tuple（顺序敏感），value 是 list 是因为同名模块在不同父节点下
      可能出现多个，导入时遇到 multi-match 要给用户提示"路径有歧义"。
    """
    by_id: dict[str, TestcaseModule] = {str(m.id): m for m in modules}

    id_to_names: dict[str, list[str]] = {}

    def _names_of(module_id: str) -> list[str]:
        if module_id in id_to_names:
            return id_to_names[module_id]
        m = by_id[module_id]
        if m.parent_id is None or str(m.parent_id) not in by_id:
            id_to_names[module_id] = [m.name]
        else:
            id_to_names[module_id] = [*_names_of(str(m.parent_id)), m.name]
        return id_to_names[module_id]

    id_to_path: dict[str | None, str] = {None: ""}
    path_to_modules: dict[tuple[str, ...], list[TestcaseModule]] = {}
    for mid, m in by_id.items():
        names = _names_of(mid)
        id_to_path[mid] = join_module_path(names)
        path_to_modules.setdefault(tuple(names), []).append(m)

    return id_to_path, path_to_modules


# ─── 内部 helpers ────────────────────────────────────────────────────


def _write_headers(ws) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF2E6BE6")
    border_side = Side(style="thin", color="FFD0D7DE")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for c, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.row_dimensions[1].height = 28
    # 表头永远可见，方便用户上下滚动时不丢上下文
    ws.freeze_panes = "A2"

    # 数据区默认换行，让多行步骤单元格直接展示完整内容
    ws.sheet_view.showGridLines = True


def _apply_data_validations(ws, *, max_row: int) -> None:
    """给 优先级 / 状态 / 执行结果 三列加 Excel 下拉约束。"""

    def _add_dv(col_key: str, options: list[str]) -> None:
        col_idx = next(
            (i for i, (_, k, _) in enumerate(COLUMNS, start=1) if k == col_key),
            None,
        )
        if col_idx is None:
            return
        formula = '"' + ",".join(options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error = "请选择下拉列表里的取值"
        dv.errorTitle = "取值非法"
        dv.prompt = "可填：" + " / ".join(options)
        dv.promptTitle = "可选项"
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{max_row}")
        ws.add_data_validation(dv)

    _add_dv("priority", PRIORITY_OPTIONS)
    _add_dv("status", STATUS_OPTIONS)
    _add_dv("exec_result", EXEC_RESULT_OPTIONS)


def _add_instructions_sheet(wb: Workbook) -> None:
    """单独一个 sheet 写字段说明，避免主表头被注释撑得过宽。"""
    if INSTRUCTIONS_SHEET_NAME in wb.sheetnames:
        return
    ws = wb.create_sheet(INSTRUCTIONS_SHEET_NAME)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 70

    rows = [
        ("列名", "必填", "说明"),
        (
            "用例编号", "否",
            "形如 TC-0001。留空 = 新增；填写但项目里查不到 = 也按新增处理；"
            "填写且项目里已存在 = 按编号匹配做覆盖式更新。",
        ),
        (
            "模块路径", "否",
            "用 / 分隔层级，例如『登录模块/账号登录』。导入时找不到的层级会自动创建。"
            "留空时归为『未分类』。",
        ),
        ("用例标题", "是", "本行唯一必填列；为空则整行被跳过。"),
        ("优先级", "否", "可填：高 / 中 / 低（也兼容 high/medium/low、p0/p1/p2）。默认『中』。"),
        ("状态", "否", "可填：有效 / 草稿 / 废弃。默认『有效』。"),
        ("执行结果", "否", "可填：未执行 / 通过 / 失败 / 阻塞。默认『未执行』。"),
        ("前置条件", "否", "多行文本；导入时按 Excel 单元格内容原样写入。"),
        (
            "步骤动作", "否",
            "多行文本，每行 = 一个步骤的操作描述。第 N 行就是第 N 个步骤。",
        ),
        (
            "预期结果", "否",
            "多行文本，每行 = 对应步骤的预期结果。"
            "若与『步骤动作』行数不一致，缺失的一边按空补齐。",
        ),
    ]
    for r, (col_a, col_b, col_c) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=col_a)
        ws.cell(row=r, column=2, value=col_b)
        ws.cell(row=r, column=3, value=col_c)
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        if r == 1:
            for c in (1, 2, 3):
                ws.cell(row=r, column=c).font = Font(bold=True)


def _pick_data_sheet(wb: Workbook):
    """优先选标准命名的工作表；否则用第一个工作表（兼容旧文件）。"""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    return wb.worksheets[0]


def _is_meaningful(v) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    return True


def _cell_str(row_values: tuple, idx: int | None) -> str:
    """把 Excel 单元格值统一成 ``str``，并去掉首尾空白。

    - 数字会被 openpyxl 读成 int/float —— ``str(int)`` 即可
    - 浮点 ``5.0`` 这种"看上去是整数"的也归一化成 ``"5"``，避免 case_no 解析掉链
    - None 返回空串
    """
    if idx is None or idx >= len(row_values):
        return ""
    v = row_values[idx]
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _parse_display_id(value: str) -> int | None:
    """``TC-0001`` / ``TC1`` / ``tc_3`` / ``42`` → ``int``；非法 → ``None``。"""
    s = value.strip()
    if not s:
        return None
    m = DISPLAY_ID_RE.match(s)
    if m:
        return int(m.group(1))
    if s.isdigit():
        return int(s)
    return None


def _normalize_value(raw: str, lookup: dict[str, str], *, default: str) -> str | None:
    """``raw`` 在 lookup 里查得到 → 返回标准 key；查不到且非空 → ``None``（让调用方报错）；
    raw 为空 → ``default``。"""
    if not raw:
        return default
    return lookup.get(raw.strip().lower()) or lookup.get(raw.strip())


def _parse_steps(action_cell: str, expected_cell: str) -> list[ParsedStep]:
    """把"步骤动作"和"预期结果"两个多行单元格对齐成步骤列表。

    Excel cell 里换行会被 openpyxl 解析成 ``\n``。空行被跳过；若两列行数对不
    上，多出来的那边继续作为独立步骤补齐——比"硬截断"对人友好。
    """
    actions = [line.rstrip() for line in (action_cell or "").splitlines() if line.strip()]
    expecteds = [line.rstrip() for line in (expected_cell or "").splitlines() if line.strip()]

    n = max(len(actions), len(expecteds))
    steps: list[ParsedStep] = []
    for i in range(n):
        action = actions[i] if i < len(actions) else ""
        expected = expecteds[i] if i < len(expecteds) else ""
        # 两边都空 → 跳过；只填了一边 → 仍当作一个步骤记录下来
        if not action and not expected:
            continue
        steps.append(ParsedStep(
            step_number=i + 1,
            action=action or "(无动作描述)",
            expected_result=expected or None,
        ))
    return steps
