"""测试用例 Excel 导入 / 导出 / 模板的纯函数单测。

只覆盖 ``excel_io.py`` 里那些不依赖 DB 的工具：模板生成、xlsx 解析、
模块路径分割与索引构造。涉及 DB 的批量导入流程（``import_testcases_xlsx``）
仍由前端集成 + 手工冒烟覆盖，保持与本仓既有 ``test_data/test_import.py``
同样的纯函数分层风格。
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field

import pytest
from openpyxl import load_workbook

from app.modules.testcases.excel_io import (
    COLUMNS,
    EXEC_RESULT_DISPLAY,
    INSTRUCTIONS_SHEET_NAME,
    PRIORITY_DISPLAY,
    SHEET_NAME,
    STATUS_DISPLAY,
    build_export_workbook,
    build_module_path_index,
    build_template_workbook,
    join_module_path,
    parse_workbook,
    split_module_path,
)
from app.modules.testcases.schemas import format_case_display_id

# ─── 模板生成 ────────────────────────────────────────────────────────


def test_template_has_expected_headers_and_sheets() -> None:
    raw = build_template_workbook(sample=False)
    wb = load_workbook(io.BytesIO(raw))
    assert SHEET_NAME in wb.sheetnames
    assert INSTRUCTIONS_SHEET_NAME in wb.sheetnames

    ws = wb[SHEET_NAME]
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(COLUMNS) + 1)]
    expected = [h for h, _, _ in COLUMNS]
    assert headers == expected


def test_template_sample_rows_render_when_requested() -> None:
    raw = build_template_workbook(sample=True)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb[SHEET_NAME]
    # row 2 是新增样例（用例编号留空），row 3 是更新样例（编号 TC-0001）
    assert ws.cell(row=2, column=1).value in (None, "")
    assert str(ws.cell(row=3, column=1).value).upper().startswith("TC-")


def test_template_priority_uses_chinese_options() -> None:
    raw = build_template_workbook(sample=True)
    wb = load_workbook(io.BytesIO(raw))
    ws = wb[SHEET_NAME]
    # 第 2 行示例的优先级应当用中文渲染
    priority_col = next(i for i, (_, k, _) in enumerate(COLUMNS, start=1) if k == "priority")
    assert ws.cell(row=2, column=priority_col).value in PRIORITY_DISPLAY.values()


# ─── 解析：基本结构 ──────────────────────────────────────────────────


def _build_xlsx(rows: list[list[object | None]]) -> bytes:
    """从 [header_row, data_rows...] 列表构造一个 xlsx 字节流，方便单测拼数据。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_parse_minimum_columns_and_defaults() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题"],
        ["", "登录成功"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert len(rows) == 1
    row = rows[0]
    assert row.title == "登录成功"
    assert row.case_no is None
    assert row.priority == "medium"
    assert row.status == "active"
    assert row.exec_result == "not_run"
    assert row.steps == []


def test_parse_full_row_with_chinese_values() -> None:
    raw = _build_xlsx([
        ["用例编号", "模块路径", "用例标题", "优先级", "状态", "执行结果",
         "前置条件", "步骤动作", "预期结果"],
        ["TC-0007", "登录/账号登录", "用正确账号登录", "高", "有效", "通过",
         "已注册账号", "打开页面\n输入账号\n点击登录", "页面 OK\n输入框 OK\n跳转首页"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert len(rows) == 1
    row = rows[0]
    assert row.case_no == 7
    assert row.module_path == "登录/账号登录"
    assert row.priority == "high"
    assert row.status == "active"
    assert row.exec_result == "passed"
    assert row.precondition == "已注册账号"
    assert [s.action for s in row.steps] == ["打开页面", "输入账号", "点击登录"]
    assert [s.expected_result for s in row.steps] == ["页面 OK", "输入框 OK", "跳转首页"]


def test_parse_accepts_english_value_aliases() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题", "优先级", "状态", "执行结果"],
        ["", "case A", "high", "deprecated", "blocked"],
        ["", "case B", "p0", "draft", "fail"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert rows[0].priority == "high"
    assert rows[0].status == "deprecated"
    assert rows[0].exec_result == "blocked"
    assert rows[1].priority == "high"
    assert rows[1].status == "draft"
    assert rows[1].exec_result == "failed"


def test_parse_skips_blank_rows() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题"],
        ["", "case A"],
        [None, None],
        ["", ""],
        ["", "case B"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert [r.title for r in rows] == ["case A", "case B"]


def test_parse_missing_title_yields_row_error() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题"],
        ["", ""],   # 真正空行 → 静默跳过
        ["TC-0001", "   "],  # 只有用例编号，标题留空白 → 行错误
    ])
    rows, errors = parse_workbook(raw)
    assert rows == []
    assert len(errors) == 1
    assert errors[0].row == 3
    assert "用例标题" in errors[0].message


def test_parse_invalid_priority_yields_row_error() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题", "优先级"],
        ["", "case A", "紧急"],
    ])
    rows, errors = parse_workbook(raw)
    assert rows == []
    assert len(errors) == 1
    assert "优先级" in errors[0].message


def test_parse_invalid_status_yields_row_error() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题", "状态"],
        ["", "case A", "归档"],
    ])
    rows, errors = parse_workbook(raw)
    assert rows == []
    assert len(errors) == 1
    assert "状态" in errors[0].message


def test_parse_invalid_display_id_records_error_but_treats_as_new() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题"],
        ["XX-9", "case A"],
    ])
    rows, errors = parse_workbook(raw)
    # 编号格式错只是 warning：行仍然以"新增"语义保留，case_no = None
    assert len(rows) == 1
    assert rows[0].case_no is None
    assert any("用例编号" in e.message for e in errors)


def test_parse_numeric_only_display_id_is_accepted() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题"],
        [42, "case A"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert rows[0].case_no == 42


def test_parse_missing_title_header_raises() -> None:
    raw = _build_xlsx([
        ["用例编号", "优先级"],
        ["", "高"],
    ])
    with pytest.raises(ValueError) as ei:
        parse_workbook(raw)
    assert "用例标题" in str(ei.value)


def test_parse_invalid_xlsx_raises() -> None:
    with pytest.raises(ValueError):
        parse_workbook(b"this is not an xlsx")


def test_parse_step_columns_uneven_lines_pad_with_empty() -> None:
    raw = _build_xlsx([
        ["用例编号", "用例标题", "步骤动作", "预期结果"],
        ["", "case A", "动作1\n动作2\n动作3", "预期1"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert len(rows[0].steps) == 3
    assert rows[0].steps[0].expected_result == "预期1"
    assert rows[0].steps[1].expected_result is None
    assert rows[0].steps[2].expected_result is None


def test_parse_header_aliases_are_recognized() -> None:
    raw = _build_xlsx([
        ["ID", "模块", "标题", "操作步骤", "预期"],
        ["TC-0002", "登录", "case A", "step 1", "expected 1"],
    ])
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert rows[0].case_no == 2
    assert rows[0].module_path == "登录"
    assert rows[0].steps[0].action == "step 1"
    assert rows[0].steps[0].expected_result == "expected 1"


# ─── 模块路径工具 ────────────────────────────────────────────────────


def test_split_module_path_handles_extra_whitespace_and_empty_segments() -> None:
    assert split_module_path("a / b /c") == ["a", "b", "c"]
    assert split_module_path("/a//b/") == ["a", "b"]
    assert split_module_path("") == []
    assert split_module_path(None) == []


def test_join_module_path_roundtrip() -> None:
    assert join_module_path(["a", "b", "c"]) == "a/b/c"
    assert join_module_path([]) == ""


@dataclass
class _FakeModule:
    """``build_module_path_index`` 只读 ``id / parent_id / name``，最小化造对象。"""

    id: uuid.UUID
    name: str
    parent_id: uuid.UUID | None = None
    project_id: uuid.UUID = field(default_factory=uuid.uuid4)


def test_build_module_path_index_renders_full_path() -> None:
    root = _FakeModule(id=uuid.uuid4(), name="登录模块")
    child = _FakeModule(id=uuid.uuid4(), name="账号登录", parent_id=root.id)
    grand = _FakeModule(id=uuid.uuid4(), name="忘记密码", parent_id=child.id)

    id_to_path, path_to_modules = build_module_path_index([root, child, grand])
    assert id_to_path[None] == ""
    assert id_to_path[str(root.id)] == "登录模块"
    assert id_to_path[str(child.id)] == "登录模块/账号登录"
    assert id_to_path[str(grand.id)] == "登录模块/账号登录/忘记密码"

    # path_to_modules 的 key 是 tuple，value 是 list（同名兄弟可能共存）
    assert path_to_modules[("登录模块", "账号登录")][0].id == child.id


def test_build_module_path_index_records_ambiguous_siblings() -> None:
    root = _FakeModule(id=uuid.uuid4(), name="支付")
    a = _FakeModule(id=uuid.uuid4(), name="退款", parent_id=root.id)
    b = _FakeModule(id=uuid.uuid4(), name="退款", parent_id=root.id)

    _, path_to_modules = build_module_path_index([root, a, b])
    assert {m.id for m in path_to_modules[("支付", "退款")]} == {a.id, b.id}


# ─── 导出 → 解析 round-trip ──────────────────────────────────────────


@dataclass
class _FakeStep:
    step_number: int
    action: str
    expected_result: str | None


@dataclass
class _FakeTestcase:
    """与 ORM 上 Testcase 对齐的最小字段集，用于走 build_export_workbook。"""

    id: uuid.UUID
    case_no: int
    module_id: uuid.UUID | None
    title: str
    precondition: str | None
    priority: str
    status: str
    exec_result: str
    steps: list[_FakeStep]


def test_export_then_parse_is_lossless_for_supported_fields() -> None:
    module_id = uuid.uuid4()
    cases = [
        _FakeTestcase(
            id=uuid.uuid4(), case_no=1, module_id=module_id,
            title="正确账号登录",
            precondition="已注册",
            priority="high", status="active", exec_result="passed",
            steps=[
                _FakeStep(1, "打开登录页", "页面 OK"),
                _FakeStep(2, "输入账号密码", "无报错"),
                _FakeStep(3, "点击登录", "跳转首页"),
            ],
        ),
        _FakeTestcase(
            id=uuid.uuid4(), case_no=2, module_id=None,
            title="错误密码登录",
            precondition=None,
            priority="medium", status="draft", exec_result="not_run",
            steps=[],
        ),
    ]
    raw = build_export_workbook(
        cases,
        lambda mid: "登录/账号登录" if mid == module_id else "",
    )
    rows, errors = parse_workbook(raw)
    assert errors == []
    assert len(rows) == 2

    by_no = {r.case_no: r for r in rows}
    assert by_no[1].title == "正确账号登录"
    assert by_no[1].priority == "high"
    assert by_no[1].status == "active"
    assert by_no[1].exec_result == "passed"
    assert by_no[1].module_path == "登录/账号登录"
    assert [s.action for s in by_no[1].steps] == ["打开登录页", "输入账号密码", "点击登录"]
    assert [s.expected_result for s in by_no[1].steps] == ["页面 OK", "无报错", "跳转首页"]

    assert by_no[2].title == "错误密码登录"
    assert by_no[2].priority == "medium"
    assert by_no[2].module_path is None
    assert by_no[2].steps == []


def test_format_display_id_padding_matches_export_value() -> None:
    # 导出 → 解析的 round-trip 依赖 ``TC-0001`` 这种 4 位补 0 编号
    assert format_case_display_id(1) == "TC-0001"
    assert format_case_display_id(123) == "TC-0123"
    assert format_case_display_id(99999) == "TC-99999"
    assert format_case_display_id(0) == "TC-?"
    assert format_case_display_id(None) == "TC-?"


# ─── 中英文枚举值的对称映射断言 ───────────────────────────────────────


def test_priority_status_exec_options_are_self_consistent() -> None:
    # 单测保护：前端 / 模板下拉 / 后端枚举三处取值必须可互查
    assert set(PRIORITY_DISPLAY) == {"high", "medium", "low"}
    assert set(STATUS_DISPLAY) == {"active", "draft", "deprecated"}
    assert set(EXEC_RESULT_DISPLAY) == {"not_run", "passed", "failed", "blocked"}
